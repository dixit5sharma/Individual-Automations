import openpyxl
from openpyxl.styles import Font,NamedStyle,Alignment           # Alignment - New learning

wb = openpyxl.Workbook()
sheet = wb[wb.sheetnames[0]]

fontObj = Font(name="Times New Roman",size=14,bold=True)
alignObj =  Alignment(horizontal="center",vertical="center")        # Value must be one of {'centerContinuous', 'general', 'distributed', 'left', 'right', 'justify', 'fill', 'center'}
styleObj = NamedStyle(name="styleObj",font=fontObj,alignment=alignObj)
sheet.cell(row=1,column=1).value = "N"
sheet.cell(row=1,column=2).value = "N+5"
sheet.cell(row=1,column=3).value = "SUM"
# print(sheet[1])               # Full Row
for each in sheet[1]:
    each.style = styleObj
for i in range(1,11):
    sheet.cell(row=i+1,column=1).value = i
    sheet.cell(row=i+1,column=2).value = i+5
    sheet.cell(row=i+1, column=3).value = "=SUM(A"+str(i+1)+":B"+str(i+1)+")"

wb.save("Files\\Formulas_1.xlsx")