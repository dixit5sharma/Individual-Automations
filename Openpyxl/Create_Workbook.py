import openpyxl

wb = openpyxl.Workbook()
print("Before setting =",wb.sheetnames)
# wb[wb.sheetnames[0]].title = "Eggs"
wb.active.title = "Eggs"
print("After setting =",wb.sheetnames)
wb.create_sheet(title="start",index=0)      # index defines the order of the sheets and title is the name
wb.create_sheet(title="end",index=len(wb.sheetnames))       # Last sheet defined

# wb.remove(wb["end"])  # Removes sheet with "end" sheet name
# del wb["end"]       # # Removes sheet with "end" sheet name

sheet = wb["start"]
sheet["A1"] = "Hello World"     # Writing values to cells is much like writing values to keys in a dictionary
wb.save("Files\\CreatedWorkbook.xlsx")