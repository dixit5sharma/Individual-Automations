import openpyxl
import time

wb = openpyxl.Workbook()
sheet = wb.active

sheet.merge_cells("A1:C3")
sheet["A1"] = "Cube 9"      # Need to set value of the top-left cell of the merged group to make it work. Else Blank

sheet.merge_cells("C5:F10")
sheet["C5"] = "I don't know what this shape is"
wb.save("Files\\Merge_UnMerge.xlsx")

time.sleep(5)

sheet.unmerge_cells("C5:F10")
wb.save("Files\\Merge_UnMerge.xlsx")