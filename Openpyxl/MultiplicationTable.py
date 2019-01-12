import openpyxl,sys
from openpyxl.styles import Font,NamedStyle

if len(sys.argv)==2:
    n = int(sys.argv[1])
    if n<=0:
        raise ValueError("Invalid Data. Number should be greater than Zero")
else:
    n = int(input("Enter N for Table\n"))
    if n<=0:
        raise ValueError("Invalid Data. Number should be greater than Zero")

wb = openpyxl.Workbook()
sheet = wb.active
fontObj = Font(size=10,bold=True)     # name="Times New Roman"
styleObj = NamedStyle(name="StyleObject",font=fontObj)

for i in range(1,n+1):
    sheet.cell(row=1,column=i+1).value = i
    sheet.cell(row=1, column=i + 1).style = styleObj
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=i + 1, column=1).style = styleObj
    for j in range(1,n+1):
        sheet.cell(row=i+1,column=j+1).value = "="+sheet.cell(row=i+1,column=j+1).column+"1*A"+str(i+1)

wb.save("Files\\MultiplicationTable.xlsx")
