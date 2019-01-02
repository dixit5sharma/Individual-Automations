import openpyxl

wb = openpyxl.load_workbook("Files\\produceSales.xlsx")
sheet = wb.active

PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

""" Yeh hai aam zindagi """
# for i in range(2,sheet.max_row+1):
#     if sheet.cell(row=i,column=1).value=="Celery":
#         sheet.cell(row=i, column=2).value = 1.19
#     if sheet.cell(row=i,column=1).value=="Garlic":
#         sheet.cell(row=i, column=2).value = 3.07
#     if sheet.cell(row=i,column=1).value=="Lemon":
#         sheet.cell(row=i, column=2).value = 1.27

""" Aur yeh hai Mentos zindagi """
for i in range(2,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value in PRICE_UPDATES:
        sheet.cell(row=i, column=2).value = PRICE_UPDATES[sheet.cell(row=i,column=1).value]

wb.save("Files\\UpdatedProduceSales.xlsx")