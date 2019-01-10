import openpyxl

wb = openpyxl.load_workbook("Files\\Formulas_1.xlsx")
sheet = wb.active

for i in range(2,sheet.max_row):
    print(sheet.cell(row=i,column=3).value, end="   ")

print()
wb1 = openpyxl.load_workbook(filename="Files\\Formulas_1.xlsx", data_only=True)     # Gives value output
sheet1 = wb1.active

for i in range(2,sheet1.max_row+1):
    # print(i,end=",")
    print(sheet1.cell(row=i,column=3).value, end="   ")

""" For data_only=True to work, kindly ensure xlsx is saved manually atleast once """
