"""
This exercise is very informative for dictionaries. Dictionaries have setdefault() function which can be used
wherever necessary and this is one example where it is put to use extraordinarily
"""

import openpyxl,pprint

wb = openpyxl.load_workbook("Files\\censuspopdata.xlsx")
sheet = wb[wb.sheetnames[0]]

# My Method 1 (Wrong)  -   Assuming each row has tract value
# print(sheet.max_row-1)

# My Method 2 (Wrong)  -   Slicing 1 column
# tract = sheet["A1":"A"+str(sheet.max_row)]
# print(len(tract)-1)

# My Method 3 (Correct)  -   Efficient since blanks/duplicates will be removed if any
# d={}
# sum={"Total":0}
# for i in range(2,sheet.max_row+1):
#     if sheet.cell(row=i,column=3).value in d:
#         d[sheet.cell(row=i,column=3).value]+=int(sheet.cell(row=i,column=4).value)
#     else:
#         d[sheet.cell(row=i, column=3).value] = int(sheet.cell(row=i, column=4).value)
#     sum["Total"]+=int(sheet.cell(row=i,column=4).value)
# pprint.pprint(d)       # Printing this data
# print("Total population =",sum["Total"])


# Book Method 4     -   Format countyData['AK']['Anchorage']['pop']
""" Yeh hai aam zindagi """
# countyData = {}
# for i in range(2,sheet.max_row+1):
#     if sheet["B"+str(i)].value in countyData:
#         if sheet["C"+str(i)].value in countyData[sheet["B"+str(i)].value]:
#             if "pop" in countyData[sheet["B"+str(i)].value][sheet["C"+str(i)].value]:
#                 countyData[sheet["B" + str(i)].value][sheet["C" + str(i)].value]["pop"] += sheet["D"+str(i)].value
#             else:
#                 countyData[sheet["B" + str(i)].value][sheet["C" + str(i)].value]["pop"] = sheet["D" + str(i)].value
#
#             if "tracts" in countyData[sheet["B" + str(i)].value][sheet["C" + str(i)].value]:
#                 countyData[sheet["B" + str(i)].value][sheet["C" + str(i)].value]["tracts"] += 1
#             else:
#                 countyData[sheet["B" + str(i)].value][sheet["C" + str(i)].value]["tracts"] = 1
#
#         else:
#             countyData[sheet["B"+str(i)].value][sheet["C"+str(i)].value] = {"pop":sheet["D"+str(i)].value,"tracts":1}
#     else:
#         countyData[sheet["B" + str(i)].value] = {sheet["C"+str(i)].value:{"pop":sheet["D"+str(i)].value,"tracts":1}}


# Method 5  -   Evem more efficient way -   Setting Defaults in Dictionary - Very useful
""" Aur yeh hai Mentos Zindagi """
countyData = {}
for i in range(2,sheet.max_row+1):
    countyData.setdefault(sheet["B"+str(i)].value,{})
    countyData[sheet["B"+str(i)].value].setdefault(sheet["C"+str(i)].value,{"pop":0,"tracts":0})
    countyData[sheet["B" + str(i)].value][sheet["C" + str(i)].value]["pop"] += sheet["D"+str(i)].value
    countyData[sheet["B" + str(i)].value][sheet["C" + str(i)].value]["tracts"] += 1

resultFile = open('Files\\census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done')