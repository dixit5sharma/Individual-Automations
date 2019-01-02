import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

wb = openpyxl.load_workbook("Files\\example.xlsx")
print("WB Sheet Names =",wb.sheetnames,";; WB Type =",type(wb),";; Active WB =",wb.active)
# OLD METHOD - wb.get_sheet_names()

Act_Sheet = wb["Sheet1"]        # OLD METHOD - wb.get_sheet_by_name("Sheet3")
print("Act_Sheet =",Act_Sheet,";; Type =",type(Act_Sheet),";; Title =",Act_Sheet.title)
Act_Cell = Act_Sheet["A1"]
print("Act_Cell =",Act_Cell,";; Type =",type(Act_Cell),";; Value =",Act_Cell.value,";; Type of value =", type(Act_Cell.value))     # datetime type)

print("column =",Act_Cell.column,";; row =",Act_Cell.row,";; coordinate =",Act_Cell.coordinate)

for counter,data in enumerate(range(1,8,2)):
    print(str(counter)+".",Act_Sheet.cell(row=data,column=2).value,end=" ;; ")

print("\nSheet Boundaries - max_row =",Act_Sheet.max_row,";; max_column =",Act_Sheet.max_column)

print("Column Letter at 100th Position =",get_column_letter(100))
print("Column Index of String 'DIX' =",column_index_from_string("DIX"))
